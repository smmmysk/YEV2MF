import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from datetime import datetime
import os
import glob
import subprocess
import sys
import shutil

# Klasör yolları
base_dir = os.path.dirname(os.path.abspath(__file__))
xml_folder = os.path.join(base_dir, "XMLYevmiye")
excel_folder = os.path.join(base_dir, "ExcelMuhasebeFisi")
excel_template = os.path.join(base_dir, "Sablon", "fis_aktarim_sablon.xlsx")

def run_xml_to_excel50():
    """Xml_to_excel50.py dosyasını çalıştır"""
    try:
        xml50_path = os.path.join(base_dir, "Xml_to_excel50.py")
        if os.path.exists(xml50_path):
            print("\n" + "="*50)
            print("İkinci aşamaya geçiliyor (Xml_to_excel50.py)...")
            print("="*50)
            
            # Mevcut Python yorumlayıcısını kullanarak diğer scripti çalıştır
            subprocess.run([sys.executable, xml50_path], check=True)
            return True
        else:
            print("Hata: Xml_to_excel50.py dosyası bulunamadı!")
            return False
    except Exception as e:
        print(f"Xml_to_excel50.py çalıştırılırken hata: {str(e)}")
        return False

def parse_xml(xml_file):
    """XML dosyasından verileri çıkarır"""
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        namespaces = {
            'gl-cor': 'http://www.xbrl.org/int/gl/cor/2006-10-25',
            'gl-bus': 'http://www.xbrl.org/int/gl/bus/2006-10-25',
            'xbrli': 'http://www.xbrl.org/2003/instance'
        }
        
        # XML'den vergi numarasını ve belge numarasını al
        tax_id = root.find('.//xbrli:identifier', namespaces)
        tax_id = tax_id.text if tax_id is not None else 'Bilinmeyen'
        
        doc_id = root.find('.//gl-cor:uniqueID', namespaces)
        doc_id = doc_id.text if doc_id is not None else ''
        
        entries = []
        
        # Tüm entryHeader'ları bul
        for entry_header in root.findall('.//gl-cor:entryHeader', namespaces):
            entry_data = {}
            
            # Fiş Numarası
            entry_number = entry_header.find('.//gl-cor:entryNumber', namespaces)
            entry_data['fis_no'] = entry_number.text if entry_number is not None else ''
            
            # Fiş Tarihi
            entry_date = entry_header.find('.//gl-cor:enteredDate', namespaces)
            if entry_date is not None and entry_date.text:
                try:
                    date_obj = datetime.strptime(entry_date.text, '%Y-%m-%d')
                    entry_data['fis_tarihi'] = date_obj.strftime('%d/%m/%Y')
                except ValueError:
                    entry_data['fis_tarihi'] = entry_date.text
            else:
                entry_data['fis_tarihi'] = ''
            
            # Fiş Açıklama
            entry_comment = entry_header.find('.//gl-cor:entryComment', namespaces)
            entry_data['fis_aciklama'] = entry_comment.text if entry_comment is not None else ''
            
            # Entry Detail'ları işle
            entry_details = []
            for detail in entry_header.findall('.//gl-cor:entryDetail', namespaces):
                detail_data = {}
                
                # Hesap Kodu
                account = detail.find('.//gl-cor:accountSubID', namespaces)
                detail_data['hesap_kodu'] = account.text if account is not None else ''
                
                # Evrak No
                doc_number = detail.find('.//gl-cor:documentNumber', namespaces)
                detail_data['evrak_no'] = doc_number.text if doc_number is not None else ''
                
                # Evrak Tarihi
                doc_date = detail.find('.//gl-cor:postingDate', namespaces)
                if doc_date is not None and doc_date.text:
                    try:
                        date_obj = datetime.strptime(doc_date.text, '%Y-%m-%d')
                        detail_data['evrak_tarihi'] = date_obj.strftime('%d/%m/%Y')
                    except ValueError:
                        detail_data['evrak_tarihi'] = doc_date.text
                else:
                    detail_data['evrak_tarihi'] = ''
                
                # Detay Açıklama
                detail_comment = detail.find('.//gl-cor:detailComment', namespaces)
                detail_data['detay_aciklama'] = detail_comment.text if detail_comment is not None else ''
                
                # Borç/Alacak
                debit_credit = detail.find('.//gl-cor:debitCreditCode', namespaces)
                amount = detail.find('.//gl-cor:amount', namespaces)
                
                if debit_credit is not None and amount is not None:
                    if debit_credit.text == 'D':
                        detail_data['borc'] = amount.text
                        detail_data['alacak'] = ''
                    elif debit_credit.text == 'C':
                        detail_data['borc'] = ''
                        detail_data['alacak'] = amount.text
                    else:
                        detail_data['borc'] = ''
                        detail_data['alacak'] = ''
                else:
                    detail_data['borc'] = ''
                    detail_data['alacak'] = ''
                
                # Belge Türü
                doc_type = detail.find('.//gl-cor:documentType', namespaces)
                doc_type_desc = detail.find('.//gl-cor:documentTypeDescription', namespaces)
                
                belge_turu = ''
                if doc_type is not None:
                    if doc_type.text == 'invoice':
                        doc_number = detail.find('.//gl-cor:documentNumber', namespaces)
                        if doc_number is not None and doc_number.text.startswith('GIB'):
                            belge_turu = 'EA'
                        else:
                            belge_turu = 'EF'
                    elif doc_type.text == 'other' and doc_type_desc is not None:
                        if doc_type_desc.text == 'Muhasebe Fişi':
                            belge_turu = 'MF'
                        elif doc_type_desc.text == 'Dekont':
                            belge_turu = 'DK'
                        elif doc_type_desc.text == 'Ücret Bordrosu İcmali':
                            belge_turu = 'ÜB'   
                        elif doc_type_desc.text == 'TAHAKKUK':
                            belge_turu = 'MF'  
                        elif doc_type_desc.text == 'Serbest Meslek Makbuzu':
                            belge_turu = 'SM'
                
                detail_data['belge_turu'] = belge_turu
                entry_details.append(detail_data)
            
            # Ana veriye ekle
            for detail in entry_details:
                entry_row = entry_data.copy()
                entry_row.update(detail)
                entries.append(entry_row)
        
        return entries
    
    except Exception as e:
        print(f"XML ayrıştırma hatası: {str(e)}")
        return None

def write_to_excel(data, excel_file, output_file):
    """Verileri Excel dosyasına yazar"""
    try:
        # Şablon dosyasını kopyala
        shutil.copy2(excel_file, output_file)
        
        # Excel dosyasını yükle
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Verileri yaz (2. satırdan başlayarak)
        for i, entry in enumerate(data, start=2):
            ws.cell(row=i, column=1).value = entry.get('fis_no', '')
            ws.cell(row=i, column=2).value = entry.get('fis_tarihi', '')
            ws.cell(row=i, column=3).value = entry.get('fis_aciklama', '')
            ws.cell(row=i, column=4).value = entry.get('hesap_kodu', '')
            ws.cell(row=i, column=5).value = entry.get('evrak_no', '')
            ws.cell(row=i, column=6).value = entry.get('evrak_tarihi', '')
            ws.cell(row=i, column=7).value = entry.get('detay_aciklama', '')
            ws.cell(row=i, column=8).value = entry.get('borc', '')
            ws.cell(row=i, column=9).value = entry.get('alacak', '')
            ws.cell(row=i, column=11).value = entry.get('belge_turu', '')
        
        # Dosyayı kaydet
        wb.save(output_file)
        return True
        
    except Exception as e:
        print(f"Excel yazma hatası: {str(e)}")
        return False

def get_output_filename(xml_file):
    """Çıktı dosyası için isim oluşturur"""
    base_name = os.path.splitext(os.path.basename(xml_file))[0]
    return os.path.join(excel_folder, f"{base_name}.xlsx")

def main():
    try:
        # Excel şablonunun varlığını kontrol et
        if not os.path.exists(excel_template):
            raise FileNotFoundError(f"Excel şablonu bulunamadı: {excel_template}")
        
        # XML dosyalarını bul
        xml_files = glob.glob(os.path.join(xml_folder, "*.xml"))
        
        if not xml_files:
            print(f"{xml_folder} klasöründe XML dosyası bulunamadı.")
            return
        
        print(f"Toplam {len(xml_files)} adet XML dosyası bulundu.")
        
        # Her XML dosyasını işle
        for xml_file in xml_files:
            try:
                print(f"\nİşleniyor: {os.path.basename(xml_file)}")
                
                # Çıktı dosya adını oluştur
                output_file = get_output_filename(xml_file)
                
                # XML'den verileri çek
                data = parse_xml(xml_file)
                if not data:
                    print(f"  Uyarı: {xml_file} dosyasından veri çekilemedi.")
                    continue
                
                # Excel'e yaz
                if write_to_excel(data, excel_template, output_file):
                    print(f"  ✓ Başarıyla oluşturuldu: {os.path.basename(output_file)}")
                else:
                    print(f"  ✗ Hata: {os.path.basename(output_file)} oluşturulamadı")
                
            except Exception as e:
                print(f"  Hata: {xml_file} işlenirken bir hata oluştu: {str(e)}")
        
        # İlk aşama tamamlandı, ikinci aşamayı başlat
        print("\nBirinci aşama tamamlandı. İkinci aşama başlatılıyor...")
        run_xml_to_excel50()
        
        print("\nTüm işlemler tamamlandı!")
        
    except Exception as e:
        print(f"Beklenmeyen bir hata oluştu: {str(e)}")

if __name__ == "__main__":
    # Gerekli klasörleri oluştur
    for folder in [xml_folder, excel_folder, os.path.dirname(excel_template)]:
        os.makedirs(folder, exist_ok=True)
    
    main()