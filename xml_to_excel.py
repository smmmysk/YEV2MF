import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from datetime import datetime
import os
import glob

# Klasör yolları
xml_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "XMLYevmiye")
excel_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ExcelMuhasebeFisi")

# Excel şablon dosya yolu
excel_template = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Sablon", "fis_aktarim_sablon.xlsx")

# XML'den verileri çekme fonksiyonu
def parse_xml(xml_file):
    try:
        # XML dosyasını yükle
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        # XML namespace'leri
        namespaces = {
            'gl-cor': 'http://www.xbrl.org/int/gl/cor/2006-10-25',
            'gl-bus': 'http://www.xbrl.org/int/gl/bus/2006-10-25',
            'xbrli': 'http://www.xbrl.org/2003/instance',
            'gl-cor': 'http://www.xbrl.org/int/gl/cor/2006-10-25',
            'xbrli': 'http://www.xbrl.org/2003/instance'
        }
        
        # Verileri saklayacak liste
        entries = []
        
        # Tüm entryHeader'ları bul
        for entry_header in root.findall('.//gl-cor:entryHeader', namespaces):
            entry_data = {}
            
            # Fiş Numarası
            entry_number = entry_header.find('.//gl-cor:entryNumber', namespaces)
            entry_data['fis_no'] = entry_number.text if entry_number is not None else ''
            
            # Fiş Tarihi
            entry_date = entry_header.find('.//gl-cor:enteredDate', namespaces)
            if entry_date is not None and entry_date.text:
                try:
                    # Tarih formatını DD/MM/YYYY'e çevir
                    date_obj = datetime.strptime(entry_date.text, '%Y-%m-%d')
                    entry_data['fis_tarihi'] = date_obj.strftime('%d/%m/%Y')
                except ValueError:
                    entry_data['fis_tarihi'] = entry_date.text
            else:
                entry_data['fis_tarihi'] = ''
            
            # Fiş Açıklama
            entry_comment = entry_header.find('.//gl-cor:entryComment', namespaces)
            entry_data['fis_aciklama'] = entry_comment.text if entry_comment is not None else ''
            
            # Entry Detail'ları işle
            entry_details = []
            for detail in entry_header.findall('.//gl-cor:entryDetail', namespaces):
                detail_data = {}
                
                # Hesap Kodu
                account = detail.find('.//gl-cor:accountSubID', namespaces)
                detail_data['hesap_kodu'] = account.text if account is not None else ''
                
                # Evrak No
                doc_number = detail.find('.//gl-cor:documentNumber', namespaces)
                detail_data['evrak_no'] = doc_number.text if doc_number is not None else ''
                
                # Evrak Tarihi
                doc_date = detail.find('.//gl-cor:postingDate', namespaces)
                if doc_date is not None and doc_date.text:
                    try:
                        # Tarih formatını DD/MM/YYYY'e çevir
                        date_obj = datetime.strptime(doc_date.text, '%Y-%m-%d')
                        detail_data['evrak_tarihi'] = date_obj.strftime('%d/%m/%Y')
                    except ValueError:
                        detail_data['evrak_tarihi'] = doc_date.text
                else:
                    detail_data['evrak_tarihi'] = ''
                
                # Detay Açıklama
                detail_comment = detail.find('.//gl-cor:detailComment', namespaces)
                detail_data['detay_aciklama'] = detail_comment.text if detail_comment is not None else ''
                
                # Borç/Alacak
                debit_credit = detail.find('.//gl-cor:debitCreditCode', namespaces)
                amount = detail.find('.//gl-cor:amount', namespaces)
                
                if debit_credit is not None and amount is not None:
                    if debit_credit.text == 'D':
                        detail_data['borc'] = amount.text
                        detail_data['alacak'] = ''
                    elif debit_credit.text == 'C':
                        detail_data['borc'] = ''
                        detail_data['alacak'] = amount.text
                    else:
                        detail_data['borc'] = ''
                        detail_data['alacak'] = ''
                else:
                    detail_data['borc'] = ''
                    detail_data['alacak'] = ''
                
                # Belge Türü
                doc_type = detail.find('.//gl-cor:documentType', namespaces)
                doc_type_desc = detail.find('.//gl-cor:documentTypeDescription', namespaces)
                
                belge_turu = ''
                if doc_type is not None:
                    if doc_type.text == 'invoice':
                        doc_number = detail.find('.//gl-cor:documentNumber', namespaces)
                        if doc_number is not None and doc_number.text.startswith('GIB'):
                            belge_turu = 'EA'
                        else:
                            belge_turu = 'EF'
                    elif doc_type.text == 'other' and doc_type_desc is not None:
                        if doc_type_desc.text == 'Muhasebe Fişi':
                            belge_turu = 'MF'
                        elif doc_type_desc.text == 'Dekont':
                            belge_turu = 'DK'
                        elif doc_type_desc.text == 'Ücret Bordrosu İcmali':
                            belge_turu = 'ÜB'   
                        elif doc_type_desc.text == 'TAHAKKUK':
                            belge_turu = 'MF'  
                        elif doc_type_desc.text == 'Serbest Meslek Makbuzu':
                            belge_turu = 'SM'  
                

                detail_data['belge_turu'] = belge_turu
                
                entry_details.append(detail_data)
            
            # Ana veriye ekle
            for detail in entry_details:
                entry_row = entry_data.copy()
                entry_row.update(detail)
                entries.append(entry_row)
        
        return entries
    
    except Exception as e:
        print(f"XML dosyası okunurken hata: {str(e)}")
        return []

# Excel'e yazma fonksiyonu
def write_to_excel(data, excel_file, output_file):
    # Excel dosyasını yükle
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Verileri yaz (2. satırdan başlayarak)
    for i, entry in enumerate(data, start=2):
        ws.cell(row=i, column=1).value = entry.get('fis_no', '')
        ws.cell(row=i, column=2).value = entry.get('fis_tarihi', '')
        ws.cell(row=i, column=3).value = entry.get('fis_aciklama', '')
        ws.cell(row=i, column=4).value = entry.get('hesap_kodu', '')
        ws.cell(row=i, column=5).value = entry.get('evrak_no', '')
        ws.cell(row=i, column=6).value = entry.get('evrak_tarihi', '')
        ws.cell(row=i, column=7).value = entry.get('detay_aciklama', '')
        ws.cell(row=i, column=8).value = entry.get('borc', '')
        ws.cell(row=i, column=9).value = entry.get('alacak', '')
        ws.cell(row=i, column=11).value = entry.get('belge_turu', '')
    
    # Dosyayı kaydet
    wb.save(output_file)
    print(f"Veriler başarıyla {output_file} dosyasına yazıldı.")

# Dosya adını oluşturma fonksiyonu
def get_output_filename(xml_file):
    try:
        # XML dosyasını oku
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        # Namespace'leri tanımla
        namespaces = {
            'xbrli': 'http://www.xbrl.org/2003/instance',
            'gl-cor': 'http://www.xbrl.org/int/gl/cor/2006-10-25'
        }
        
        # Vergi numarasını al
        tax_id = root.find('.//xbrli:identifier', namespaces)
        tax_id = tax_id.text if tax_id is not None else 'Bilinmeyen'
        
        # Belge numarasını al (YEV202501000001 gibi)
        doc_id = root.find('.//gl-cor:uniqueID', namespaces)
        doc_id = doc_id.text if doc_id is not None else ''
        
        # Dosya adını oluştur (örn: 22462442242_YEV202501000001.xlsx)
        filename = f"{tax_id}_{doc_id}.xlsx"
        return os.path.join(excel_folder, filename)
    except Exception as e:
        print(f"Dosya adı oluşturulurken hata: {str(e)}")
        return os.path.join(excel_folder, f"output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

# Ana işlem
def main():
    try:
        # Excel şablonunun varlığını kontrol et
        if not os.path.exists(excel_template):
            raise FileNotFoundError(f"Excel şablonu bulunamadı: {excel_template}")
        
        # XML dosyalarını bul
        xml_files = glob.glob(os.path.join(xml_folder, "*.xml"))
        
        if not xml_files:
            print(f"{xml_folder} klasöründe XML dosyası bulunamadı.")
            return
        
        print(f"Toplam {len(xml_files)} adet XML dosyası bulundu.")
        
        # Her XML dosyasını işle
        for xml_file in xml_files:
            try:
                print(f"\nİşleniyor: {os.path.basename(xml_file)}")
                
                # Çıktı dosya adını oluştur
                output_file = get_output_filename(xml_file)
                
                # XML'den verileri çek
                data = parse_xml(xml_file)
                
                # Excel'e yaz
                write_to_excel(data, excel_template, output_file)
                
                print(f"Başarıyla oluşturuldu: {os.path.basename(output_file)}")
                
            except Exception as e:
                print(f"Hata: {xml_file} işlenirken bir hata oluştu: {str(e)}")
        
        print("\nTüm işlemler tamamlandı!")
        
    except Exception as e:
        print(f"Beklenmeyen bir hata oluştu: {str(e)}")

if __name__ == "__main__":
    main()
