import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from datetime import datetime
import os
import glob
import shutil

# Klasör yolları
xml_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "XMLYevmiye")
excel_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ExcelMuhasebeFisi")

# Excel şablon dosya yolu
excel_template = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Sablon", "fis_aktarim_sablon.xlsx")

def parse_xml(xml_file):
    try:
        # XML dosyasını yükle
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        # XML namespace'leri
        namespaces = {
            'gl-cor': 'http://www.xbrl.org/int/gl/cor/2006-10-25',
            'gl-bus': 'http://www.xbrl.org/int/gl/bus/2006-10-25',
            'xbrli': 'http://www.xbrl.org/2003/instance'
        }
        
        # XML'den vergi numarasını ve belge numarasını al
        tax_id = root.find('.//xbrli:identifier', namespaces)
        tax_id = tax_id.text if tax_id is not None else 'Bilinmeyen'
        
        doc_id = root.find('.//gl-cor:uniqueID', namespaces)
        doc_id = doc_id.text if doc_id is not None else ''
        
        # Fişleri saklamak için sözlük
        fis_dict = {}
        
        # Tüm entryHeader'ları bul
        for entry_header in root.findall('.//gl-cor:entryHeader', namespaces):
            # Fiş Numarası
            entry_number = entry_header.find('.//gl-cor:entryNumber', namespaces)
            fis_no = entry_number.text if entry_number is not None else ''
            
            # Fiş Tarihi
            entry_date = entry_header.find('.//gl-cor:enteredDate', namespaces)
            if entry_date is not None and entry_date.text:
                try:
                    date_obj = datetime.strptime(entry_date.text, '%Y-%m-%d')
                    fis_tarihi = date_obj.strftime('%d/%m/%Y')
                except ValueError:
                    fis_tarihi = entry_date.text
            else:
                fis_tarihi = ''
            
            # Fiş Açıklama
            entry_comment = entry_header.find('.//gl-cor:entryComment', namespaces)
            fis_aciklama = entry_comment.text if entry_comment is not None else ''
            
            # Fişi oluştur veya güncelle
            if fis_no not in fis_dict:
                fis_dict[fis_no] = {
                    'fis_no': fis_no,
                    'fis_tarihi': fis_tarihi,
                    'fis_aciklama': fis_aciklama,
                    'kalemler': []
                }
            
            # Entry Detail'ları işle
            for detail in entry_header.findall('.//gl-cor:entryDetail', namespaces):
                detail_data = {}
                
                # Hesap Kodu
                account = detail.find('.//gl-cor:accountSubID', namespaces)
                detail_data['hesap_kodu'] = account.text if account is not None else ''
                
                # Evrak No
                doc_number = detail.find('.//gl-cor:documentNumber', namespaces)
                detail_data['evrak_no'] = doc_number.text if doc_number is not None else ''
                
                # Evrak Tarihi
                doc_date = detail.find('.//gl-cor:postingDate', namespaces)
                if doc_date is not None and doc_date.text:
                    try:
                        date_obj = datetime.strptime(doc_date.text, '%Y-%m-%d')
                        detail_data['evrak_tarihi'] = date_obj.strftime('%d/%m/%Y')
                    except ValueError:
                        detail_data['evrak_tarihi'] = doc_date.text
                else:
                    detail_data['evrak_tarihi'] = ''
                
                # Detay Açıklama
                detail_comment = detail.find('.//gl-cor:detailComment', namespaces)
                detail_data['detay_aciklama'] = detail_comment.text if detail_comment is not None else ''
                
                # Borç/Alacak
                debit_credit = detail.find('.//gl-cor:debitCreditCode', namespaces)
                amount = detail.find('.//gl-cor:amount', namespaces)
                
                if debit_credit is not None and amount is not None:
                    if debit_credit.text == 'D':
                        detail_data['borc'] = amount.text
                        detail_data['alacak'] = ''
                    elif debit_credit.text == 'C':
                        detail_data['borc'] = ''
                        detail_data['alacak'] = amount.text
                    else:
                        detail_data['borc'] = ''
                        detail_data['alacak'] = ''
                else:
                    detail_data['borc'] = ''
                    detail_data['alacak'] = ''
                
                # Belge Türü
                doc_type = detail.find('.//gl-cor:documentType', namespaces)
                doc_type_desc = detail.find('.//gl-cor:documentTypeDescription', namespaces)
                
                belge_turu = ''
                if doc_type is not None:
                    if doc_type.text == 'invoice':
                        doc_number = detail.find('.//gl-cor:documentNumber', namespaces)
                        if doc_number is not None and doc_number.text.startswith('GIB'):
                            belge_turu = 'EA'
                        else:
                            belge_turu = 'EF'
                    elif doc_type.text == 'other' and doc_type_desc is not None:
                        if doc_type_desc.text == 'Muhasebe Fişi':
                            belge_turu = 'MF'
                        elif doc_type_desc.text == 'Dekont':
                            belge_turu = 'DK'
                        elif doc_type_desc.text == 'Ücret Bordrosu İcmali':
                            belge_turu = 'ÜB'   
                        elif doc_type_desc.text == 'TAHAKKUK':
                            belge_turu = 'MF'  
                        elif doc_type_desc.text == 'Serbest Meslek Makbuzu':
                            belge_turu = 'SM'
                
                detail_data['belge_turu'] = belge_turu
                
                # Fişe kalemi ekle
                fis_dict[fis_no]['kalemler'].append(detail_data)
        
        # Fiş numaralarını sırala
        fis_nolari = sorted(fis_dict.keys())
        
        # Her 50 fiş için ayrı bir Excel dosyası oluştur
        for i in range(0, len(fis_nolari), 50):
            # Dosya adını oluştur
            start = i + 1
            end = min(i + 50, len(fis_nolari))
            output_file = os.path.join(
                excel_folder,
                f"{tax_id}_{doc_id}_{start}-{end}.xlsx"
            )
            
            # Şablon dosyasını kopyala
            shutil.copy2(excel_template, output_file)
            
            # Excel dosyasını yükle
            wb = load_workbook(output_file)
            ws = wb.active
            
            # Satır sayacı (başlık satırından sonra başla)
            row_num = 2
            
            # Bu partisyondaki fişleri yaz
            for fis_no in fis_nolari[i:end]:
                fis = fis_dict[fis_no]
                for kalem in fis['kalemler']:
                    ws.cell(row=row_num, column=1).value = fis['fis_no']
                    ws.cell(row=row_num, column=2).value = fis['fis_tarihi']
                    ws.cell(row=row_num, column=3).value = fis['fis_aciklama']
                    ws.cell(row=row_num, column=4).value = kalem.get('hesap_kodu', '')
                    ws.cell(row=row_num, column=5).value = kalem.get('evrak_no', '')
                    ws.cell(row=row_num, column=6).value = kalem.get('evrak_tarihi', '')
                    ws.cell(row=row_num, column=7).value = kalem.get('detay_aciklama', '')
                    ws.cell(row=row_num, column=8).value = kalem.get('borc', '')
                    ws.cell(row=row_num, column=9).value = kalem.get('alacak', '')
                    ws.cell(row=row_num, column=11).value = kalem.get('belge_turu', '')
                    row_num += 1
            
            # Dosyayı kaydet
            wb.save(output_file)
            print(f"Oluşturuldu: {os.path.basename(output_file)}")
        
        return True
        
    except Exception as e:
        print(f"Hata oluştu: {str(e)}")
        return False

def main():
    try:
        # Excel şablonunun varlığını kontrol et
        if not os.path.exists(excel_template):
            raise FileNotFoundError(f"Excel şablonu bulunamadı: {excel_template}")
        
        # XML dosyalarını bul
        xml_files = glob.glob(os.path.join(xml_folder, "*.xml"))
        
        if not xml_files:
            print(f"{xml_folder} klasöründe XML dosyası bulunamadı.")
            return
        
        print(f"Toplam {len(xml_files)} adet XML dosyası bulundu.")
        
        # Her XML dosyasını işle
        for xml_file in xml_files:
            print(f"\nİşleniyor: {os.path.basename(xml_file)}")
            parse_xml(xml_file)
        
        print("\nTüm işlemler tamamlandı!")
        
    except Exception as e:
        print(f"Beklenmeyen bir hata oluştu: {str(e)}")

if __name__ == "__main__":
    main()