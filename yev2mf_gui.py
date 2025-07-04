# yev2mf_gui.py
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import os
import sys
import subprocess
from datetime import datetime
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import shutil
from pathlib import Path
import webbrowser

class Yev2MFApp:
    def __init__(self, root):
        self.root = root
        self.root.title("YEV2MF - XML'den Excel'e Dönüşüm")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)
        
        # Stil ayarları
        self.setup_styles()
        
        # Ana frame
        self.main_frame = ttk.Frame(root, padding="15")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Başlık
        self.header_frame = ttk.Frame(self.main_frame)
        self.header_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(
            self.header_frame,
            text="YEV2MF - XML'den Excel'e Dönüşüm Aracı",
            font=('Segoe UI', 16, 'bold'),
            foreground='#2c3e50'
        ).pack(side=tk.LEFT)
        
        # Sürüm bilgisi
        ttk.Label(
            self.header_frame,
            text="v1.0.0",
            font=('Segoe UI', 8),
            foreground='#7f8c8d'
        ).pack(side=tk.RIGHT, padx=10)
        
        # Bilgi paneli
        self.setup_info_panel()
        
        # İlerleme çubuğu
        self.setup_progress_section()
        
        # Log alanı
        self.setup_log_section()
        
        # Alt bilgi
        self.setup_footer()
        
        # Klasörleri kontrol et
        self.setup_folders()
        
        # İlk çalıştırmada kontrol yap
        self.check_initial_conditions()
    
    def setup_styles(self):
        """Arayüz stillerini ayarla"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Genel stiller
        style.configure('.', font=('Segoe UI', 10))
        
        # Buton stilleri
        style.configure('TButton', padding=6, relief='flat', background='#3498db', foreground='white')
        style.map('TButton',
                 background=[('active', '#2980b9')],
                 foreground=[('active', 'white')])
        
        # Giriş alanı stilleri
        style.configure('TEntry', padding=5)
        
        # Etiket stilleri
        style.configure('TLabel', background='#ecf0f1')
        
        # Çerçeve stilleri
        style.configure('TFrame', background='#ecf0f1')
        
        # Sekme stilleri
        style.configure('TNotebook', background='#ecf0f1')
        style.configure('TNotebook.Tab', padding=[10, 5], font=('Segoe UI', 10, 'bold'))
    
    def setup_info_panel(self):
        """Bilgi panelini oluştur"""
        info_frame = ttk.LabelFrame(
            self.main_frame,
            text=" Bilgi ve Talimatlar ",
            padding=15,
            style='TFrame'
        )
        info_frame.pack(fill=tk.X, pady=(0, 15))
        
        info_text = (
            "1. XML dosyalarınızı 'XMLYevmiye' klasörüne kopyalayın\n"
            "2. 'Dönüşümü Başlat' butonuna tıklayın\n"
            "3. Dönüştürülen dosyalar 'ExcelMuhasebeFisi' klasörüne kaydedilecektir\n"
            "4. İşlem tamamlandığında 'Excel Klasörünü Aç' butonu ile sonuçlara ulaşabilirsiniz"
        )
        
        ttk.Label(
            info_frame,
            text=info_text,
            justify=tk.LEFT,
            font=('Segoe UI', 10),
            background='#ecf0f1'
        ).pack(anchor='w')
    
    def setup_progress_section(self):
        """İlerleme bölümünü oluştur"""
        progress_frame = ttk.Frame(self.main_frame)
        progress_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Butonlar
        btn_frame = ttk.Frame(progress_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.start_btn = ttk.Button(
            btn_frame,
            text="Dönüşümü Başlat",
            command=self.start_conversion,
            style='TButton'
        )
        self.start_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(
            btn_frame,
            text="XML Klasörünü Aç",
            command=self.open_xml_folder
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            btn_frame,
            text="Excel Klasörünü Aç",
            command=self.open_excel_folder
        ).pack(side=tk.LEFT, padx=5)
        
        # İlerleme çubuğu
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(
            progress_frame,
            orient=tk.HORIZONTAL,
            length=100,
            mode='determinate',
            variable=self.progress_var
        )
        self.progress.pack(fill=tk.X, pady=(5, 0))
        
        # Durum etiketi
        self.status_var = tk.StringVar(value="Hazır")
        ttk.Label(
            progress_frame,
            textvariable=self.status_var,
            font=('Segoe UI', 9),
            foreground='#2c3e50'
        ).pack(pady=(5, 0))
    
    def setup_log_section(self):
        """Log bölümünü oluştur"""
        log_frame = ttk.LabelFrame(
            self.main_frame,
            text=" İşlem Geçmişi ",
            padding=10
        )
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        # Log alanı
        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            wrap=tk.WORD,
            font=('Consolas', 9),
            bg='#2c3e50',
            fg='#ecf0f1',
            insertbackground='white',
            selectbackground='#3498db'
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # Log renklendirme etiketleri
        self.log_text.tag_config('info', foreground='#bdc3c7')
        self.log_text.tag_config('success', foreground='#2ecc71')
        self.log_text.tag_config('warning', foreground='#f39c12')
        self.log_text.tag_config('error', foreground='#e74c3c')
        self.log_text.tag_config('highlight', foreground='#3498db')
    
    def setup_footer(self):
        """Alt bilgi bölümünü oluştur"""
        footer_frame = ttk.Frame(self.main_frame)
        footer_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Label(
            footer_frame,
            text="© 2025 YEV2MF - Tüm hakları saklıdır",
            font=('Segoe UI', 8),
            foreground='#7f8c8d'
        ).pack(side=tk.LEFT)
        
        ttk.Label(
            footer_frame,
            text="Geliştirici: SMMM YAZILIM",
            font=('Segoe UI', 8, 'bold'),
            foreground='#3498db',
            cursor='hand2'
        ).pack(side=tk.RIGHT)
    
    def setup_folders(self):
        """Gerekli klasörleri oluştur"""
        self.base_dir = Path(__file__).parent
        self.xml_folder = self.base_dir / "XMLYevmiye"
        self.excel_folder = self.base_dir / "ExcelMuhasebeFisi"
        self.template_folder = self.base_dir / "Sablon"
        self.template_file = self.template_folder / "fis_aktarim_sablon.xlsx"
        
        # Klasörleri oluştur
        for folder in [self.xml_folder, self.excel_folder, self.template_folder]:
            folder.mkdir(exist_ok=True)
        
        self.log("Uygulama başlatıldı.", 'info')
        self.log(f"XML Klasörü: {self.xml_folder}", 'info')
        self.log(f"Excel Klasörü: {self.excel_folder}", 'info')
    
    def check_initial_conditions(self):
        """Başlangıç kontrollerini yap"""
        # Şablon dosyasını kontrol et
        if not self.template_file.exists():
            self.log("UYARI: Excel şablon dosyası bulunamadı!", 'warning')
            self.log(f"Lütfen '{self.template_file}' dosyasını oluşturun.", 'warning')
            self.start_btn.config(state='disabled')
        
        # XML dosyalarını kontrol et
        xml_files = list(self.xml_folder.glob("*.xml"))
        if not xml_files:
            self.log("UYARI: İşlenecek XML dosyası bulunamadı.", 'warning')
            self.log(f"Lütfen XML dosyalarınızı '{self.xml_folder}' klasörüne ekleyin.", 'warning')
    
    def log(self, message, level='info'):
        """Log mesajı ekler"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n", level)
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def open_xml_folder(self):
        """XML klasörünü açar"""
        os.startfile(str(self.xml_folder))
    
    def open_excel_folder(self):
        """Excel klasörünü açar"""
        os.startfile(str(self.excel_folder))
    
    def start_conversion(self):
        """Dönüşüm işlemini başlat"""
        # Klasörleri kontrol et
        if not self.template_file.exists():
            messagebox.showerror(
                "Hata", 
                "Excel şablon dosyası bulunamadı!\n"
                f"Lütfen '{self.template_file}' dosyasını oluşturun."
            )
            return
        
        # XML dosyalarını kontrol et
        xml_files = list(self.xml_folder.glob("*.xml"))
        if not xml_files:
            messagebox.showinfo(
                "Bilgi", 
                "İşlenecek XML dosyası bulunamadı.\n"
                f"Lütfen XML dosyalarınızı '{self.xml_folder}' klasörüne ekleyin."
            )
            return
        
        # İşlemi başlat
        self.start_btn.config(state='disabled')
        self.progress_var.set(0)
        self.status_var.set("Dönüşüm başlatılıyor...")
        self.log("\nDönüşüm başlatılıyor...", 'highlight')
        
        # Dönüşümü ayrı bir thread'de başlat
        import threading
        thread = threading.Thread(target=self.run_conversion, args=(xml_files,), daemon=True)
        thread.start()
    
    def run_conversion(self, xml_files):
        """Dönüşüm işlemini çalıştır"""
        try:
            total_files = len(xml_files)
            self.log(f"Toplam {total_files} adet XML dosyası bulundu.", 'info')
            
            for i, xml_file in enumerate(xml_files, 1):
                self.status_var.set(f"İşleniyor: {xml_file.name} ({i}/{total_files})")
                self.progress_var.set((i / total_files) * 100)
                self.log(f"\n{i}. Dosya işleniyor: {xml_file.name}", 'info')
                
                try:
                    # XML'den verileri çek
                    data = self.parse_xml(xml_file)
                    if data:
                        # Excel'e yaz
                        output_file = self.excel_folder / f"{xml_file.stem}.xlsx"
                        if self.write_to_excel(data, output_file):
                            self.log(f"  ✓ Başarıyla dönüştürüldü: {output_file.name}", 'success')
                        else:
                            self.log(f"  ✗ Dönüştürme başarısız: {xml_file.name}", 'error')
                    else:
                        self.log(f"  ✗ Dönüştürme başarısız (veri yok): {xml_file.name}", 'warning')
                
                except Exception as e:
                    self.log(f"  ✗ Hata: {str(e)}", 'error')
                
                self.root.update()
            
            self.status_var.set("Dönüşüm tamamlandı!")
            self.log("\nTüm işlemler başarıyla tamamlandı!", 'success')
            
            # İkinci aşamayı başlat
            self.log("\nİkinci aşama başlatılıyor (Xml_to_excel50.py)...", 'highlight')
            self.run_second_stage()
            
        except Exception as e:
            self.log(f"Beklenmeyen hata: {str(e)}", 'error')
            messagebox.showerror("Hata", f"İşlem sırasında bir hata oluştu:\n{str(e)}")
        
        finally:
            self.start_btn.config(state='normal')
            self.progress_var.set(0)
    
    def run_second_stage(self):
        """İkinci aşamayı çalıştır"""
        try:
            xml50_path = self.base_dir / "Xml_to_excel50.py"
            if xml50_path.exists():
                self.log("Xml_to_excel50.py çalıştırılıyor...", 'info')
                result = subprocess.run(
                    [sys.executable, str(xml50_path)],
                    capture_output=True,
                    text=True
                )
                
                if result.returncode == 0:
                    self.log("İkinci aşama başarıyla tamamlandı.", 'success')
                    if result.stdout:
                        self.log("Çıktı:\n" + result.stdout, 'info')
                else:
                    self.log(f"İkinci aşama hatası (Kod: {result.returncode}):", 'error')
                    if result.stderr:
                        self.log(result.stderr, 'error')
            else:
                self.log("UYARI: Xml_to_excel50.py dosyası bulunamadı!", 'warning')
                
        except Exception as e:
            self.log(f"İkinci aşama çalıştırılırken hata: {str(e)}", 'error')
    
    def parse_xml(self, xml_file):
        """XML dosyasını ayrıştır"""
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            
            namespaces = {
                'gl-cor': 'http://www.xbrl.org/int/gl/cor/2006-10-25',
                'gl-bus': 'http://www.xbrl.org/int/gl/bus/2006-10-25',
                'xbrli': 'http://www.xbrl.org/2003/instance'
            }
            
            # XML'den vergi numarasını ve belge numarasını al
            tax_id = root.find('.//xbrli:identifier', namespaces)
            tax_id = tax_id.text if tax_id is not None else 'Bilinmeyen'
            
            doc_id = root.find('.//gl-cor:uniqueID', namespaces)
            doc_id = doc_id.text if doc_id is not None else ''
            
            entries = []
            
            # Tüm entryHeader'ları bul
            for entry_header in root.findall('.//gl-cor:entryHeader', namespaces):
                entry_data = {}
                
                # Fiş Numarası
                entry_number = entry_header.find('.//gl-cor:entryNumber', namespaces)
                entry_data['fis_no'] = entry_number.text if entry_number is not None else ''
                
                # Fiş Tarihi
                entry_date = entry_header.find('.//gl-cor:enteredDate', namespaces)
                if entry_date is not None and entry_date.text:
                    try:
                        date_obj = datetime.strptime(entry_date.text, '%Y-%m-%d')
                        entry_data['fis_tarihi'] = date_obj.strftime('%d/%m/%Y')
                    except ValueError:
                        entry_data['fis_tarihi'] = entry_date.text
                else:
                    entry_data['fis_tarihi'] = ''
                
                # Fiş Açıklama
                entry_comment = entry_header.find('.//gl-cor:entryComment', namespaces)
                entry_data['fis_aciklama'] = entry_comment.text if entry_comment is not None else ''
                
                # Entry Detail'ları işle
                entry_details = []
                for detail in entry_header.findall('.//gl-cor:entryDetail', namespaces):
                    detail_data = {}
                    
                    # Hesap Kodu
                    account = detail.find('.//gl-cor:accountSubID', namespaces)
                    detail_data['hesap_kodu'] = account.text if account is not None else ''
                    
                    # Evrak No
                    doc_number = detail.find('.//gl-cor:documentNumber', namespaces)
                    detail_data['evrak_no'] = doc_number.text if doc_number is not None else ''
                    
                    # Evrak Tarihi
                    doc_date = detail.find('.//gl-cor:postingDate', namespaces)
                    if doc_date is not None and doc_date.text:
                        try:
                            date_obj = datetime.strptime(doc_date.text, '%Y-%m-%d')
                            detail_data['evrak_tarihi'] = date_obj.strftime('%d/%m/%Y')
                        except ValueError:
                            detail_data['evrak_tarihi'] = doc_date.text
                    else:
                        detail_data['evrak_tarihi'] = ''
                    
                    # Detay Açıklama
                    detail_comment = detail.find('.//gl-cor:detailComment', namespaces)
                    detail_data['detay_aciklama'] = detail_comment.text if detail_comment is not None else ''
                    
                    # Borç/Alacak
                    debit_credit = detail.find('.//gl-cor:debitCreditCode', namespaces)
                    amount = detail.find('.//gl-cor:amount', namespaces)
                    
                    if debit_credit is not None and amount is not None:
                        if debit_credit.text == 'D':
                            detail_data['borc'] = amount.text
                            detail_data['alacak'] = ''
                        elif debit_credit.text == 'C':
                            detail_data['borc'] = ''
                            detail_data['alacak'] = amount.text
                        else:
                            detail_data['borc'] = ''
                            detail_data['alacak'] = ''
                    else:
                        detail_data['borc'] = ''
                        detail_data['alacak'] = ''
                    
                    # Belge Türü
                    doc_type = detail.find('.//gl-cor:documentType', namespaces)
                    doc_type_desc = detail.find('.//gl-cor:documentTypeDescription', namespaces)
                    
                    belge_turu = ''
                    if doc_type is not None:
                        if doc_type.text == 'invoice':
                            doc_number = detail.find('.//gl-cor:documentNumber', namespaces)
                            if doc_number is not None and doc_number.text.startswith('GIB'):
                                belge_turu = 'EA'
                            else:
                                belge_turu = 'EF'
                        elif doc_type.text == 'other' and doc_type_desc is not None:
                            if doc_type_desc.text == 'Muhasebe Fişi':
                                belge_turu = 'MF'
                            elif doc_type_desc.text == 'Dekont':
                                belge_turu = 'DK'
                            elif doc_type_desc.text == 'Ücret Bordrosu İcmali':
                                belge_turu = 'ÜB'   
                            elif doc_type_desc.text == 'TAHAKKUK':
                                belge_turu = 'MF'  
                            elif doc_type_desc.text == 'Serbest Meslek Makbuzu':
                                belge_turu = 'SM'
                    
                    detail_data['belge_turu'] = belge_turu
                    entry_details.append(detail_data)
                
                # Ana veriye ekle
                for detail in entry_details:
                    entry_row = entry_data.copy()
                    entry_row.update(detail)
                    entries.append(entry_row)
            
            return entries
        
        except Exception as e:
            self.log(f"XML ayrıştırma hatası: {str(e)}", 'error')
            return None
    
    def write_to_excel(self, data, output_file):
        """Verileri Excel dosyasına yazar"""
        try:
            # Şablon dosyasını kopyala
            shutil.copy2(self.template_file, output_file)
            
            # Excel dosyasını yükle
            wb = load_workbook(output_file)
            ws = wb.active
            
            # Verileri yaz (2. satırdan başlayarak)
            for i, entry in enumerate(data, start=2):
                ws.cell(row=i, column=1).value = entry.get('fis_no', '')
                ws.cell(row=i, column=2).value = entry.get('fis_tarihi', '')
                ws.cell(row=i, column=3).value = entry.get('fis_aciklama', '')
                ws.cell(row=i, column=4).value = entry.get('hesap_kodu', '')
                ws.cell(row=i, column=5).value = entry.get('evrak_no', '')
                ws.cell(row=i, column=6).value = entry.get('evrak_tarihi', '')
                ws.cell(row=i, column=7).value = entry.get('detay_aciklama', '')
                ws.cell(row=i, column=8).value = entry.get('borc', '')
                ws.cell(row=i, column=9).value = entry.get('alacak', '')
                ws.cell(row=i, column=11).value = entry.get('belge_turu', '')
            
            # Dosyayı kaydet
            wb.save(output_file)
            return True
            
        except Exception as e:
            self.log(f"Excel yazma hatası: {str(e)}", 'error')
            return False

def main():
    # Ana pencereyi oluştur
    root = tk.Tk()
    
    # Pencereyi ekranın ortasında aç
    window_width = 900
    window_height = 700
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)
    root.geometry(f'{window_width}x{window_height}+{x}+{y}')
    
    # Uygulamayı başlat
    app = Yev2MFApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()